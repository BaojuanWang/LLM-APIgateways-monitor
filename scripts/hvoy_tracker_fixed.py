"""
hvoy.ai 平台追踪脚本
运行位置: scripts/hvoy_tracker.py
"""

import json
import re
import requests
import csv
import time
from datetime import datetime, timezone
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE_DIR    = Path(__file__).parent.parent
DATA_DIR    = BASE_DIR / "data"
HVOY_DIR    = DATA_DIR / "hvoy_raw"
MANUAL_CSV  = DATA_DIR / "manual_sites.csv"
FETCH_STATUS_JSON = DATA_DIR / "hvoy_fetch_status.json"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
}

COLUMNS = [
    ("rankPosition",      "排名"),
    ("siteName",          "平台名称"),
    ("siteDomain",        "域名"),
    ("verificationType",  "认证类型"),
    ("overallScore",      "综合评分"),
    ("avgOnlineRate",     "平均在线率(%)"),
    ("avgLatencyS",       "平均延迟(s)"),
    ("latencySuspicious", "延迟异常"),
    ("modelCount",        "模型数量"),
    ("rankedModelCount",  "参与排名模型数"),
    ("averageRating",     "用户评分"),
    ("reviewCount",       "评价数"),
    ("voteUpCount",       "点赞数"),
    ("voteDownCount",     "踩数"),
    ("siteDescription",   "简介"),
    ("relaySiteId",       "站点ID"),
]


def fetch_payload():
    print("正在请求 hvoy.ai ...")
    last_error = None
    for attempt in range(1, 4):
        try:
            resp = requests.get("https://www.hvoy.ai/en/sites", headers=HEADERS, timeout=30)
            resp.raise_for_status()
            break
        except requests.HTTPError as exc:
            last_error = exc
            # A 403 is an explicit refusal. Retrying it only adds unnecessary traffic.
            if exc.response is not None and exc.response.status_code == 403:
                raise
        except requests.RequestException as exc:
            last_error = exc

        if attempt < 3:
            delay = attempt * 5
            print(f"请求失败，{delay} 秒后重试（{attempt}/3）: {last_error}")
            time.sleep(delay)
    else:
        raise last_error

    match = re.search(
        r'<script[^>]+id="__RELAY_SITE_RANKINGS_PAYLOAD__"[^>]*>(.*?)</script>',
        resp.text, re.DOTALL
    )
    if not match:
        raise ValueError("找不到 __RELAY_SITE_RANKINGS_PAYLOAD__，页面结构可能已变")
    data = json.loads(match.group(1))
    items = data.get("items", [])
    print(f"抓到 {len(items)} 个平台")
    return items, data.get("updatedAt", "")


def save_fetch_status(status, item_count=0, error="", used_cache=False):
    payload = {
        "checked_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "status": status,
        "item_count": item_count,
        "used_cache": used_cache,
        "error": str(error)[:500],
    }
    with open(FETCH_STATUS_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def load_last(path):
    if not path.exists():
        return {}
    with open(path, encoding="utf-8") as f:
        items = json.load(f)
    return {item["siteDomain"]: item for item in items}


def compare(old, new_items):
    new_map = {item["siteDomain"]: item for item in new_items}
    added   = [v for k, v in new_map.items() if k not in old]
    removed = [v for k, v in old.items() if k not in new_map]
    return added, removed


def save_json(items, path):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(items, f, ensure_ascii=False, indent=2)


def save_csv(items, path):
    keys = [col[0] for col in COLUMNS]
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=["domain", "platform_name"] + keys)
        writer.writeheader()
        for item in items:
            row = {"domain": item.get("siteDomain", ""), "platform_name": item.get("siteName", "")}
            for k in keys:
                val = item.get(k)
                if isinstance(val, bool):
                    val = "是" if val else "否"
                row[k] = val
            writer.writerow(row)


def save_excel(items, added, removed, path, updated_at, date_str):
    wb = Workbook()
    header_fill  = PatternFill("solid", start_color="2D5EA2")
    header_font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    new_fill     = PatternFill("solid", start_color="D4EFDF")
    removed_fill = PatternFill("solid", start_color="FADBD8")
    alt_fill     = PatternFill("solid", start_color="F0F4FA")
    headers = [col[1] for col in COLUMNS]
    keys    = [col[0] for col in COLUMNS]
    col_widths = [6, 20, 28, 10, 10, 14, 12, 10, 10, 14, 10, 8, 8, 8, 50, 10]
    new_domains = {item.get("siteDomain") for item in added}

    def write_sheet(ws, data, highlight_fill=None):
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row_idx, item in enumerate(data, 2):
            is_new = item.get("siteDomain") in new_domains
            for col_idx, key in enumerate(keys, 1):
                val = item.get(key)
                if isinstance(val, bool):
                    val = "是" if val else "否"
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(vertical="center", wrap_text=(key == "siteDescription"))
                if highlight_fill:
                    cell.fill = highlight_fill
                elif is_new:
                    cell.fill = new_fill
                elif row_idx % 2 == 0:
                    cell.fill = alt_fill
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"

    ws1 = wb.active
    ws1.title = "全量平台"
    write_sheet(ws1, items)

    ws2 = wb.create_sheet("新增平台")
    write_sheet(ws2, added)

    ws3 = wb.create_sheet("消失平台")
    write_sheet(ws3, removed, highlight_fill=removed_fill)

    ws4 = wb.create_sheet("说明")
    meta = [
        ("数据来源",     "https://www.hvoy.ai/en/sites"),
        ("抓取时间",     date_str),
        ("hvoy更新时间", updated_at),
        ("总平台数",     len(items)),
        ("本次新增",     len(added)),
        ("本次消失",     len(removed)),
        ("说明",         "新增平台在「全量平台」sheet中以绿色高亮显示"),
    ]
    for r, (k, v) in enumerate(meta, 1):
        ws4.cell(row=r, column=1, value=k).font = Font(bold=True, name="Arial")
        ws4.cell(row=r, column=2, value=str(v)).font = Font(name="Arial")
    ws4.column_dimensions["A"].width = 16
    ws4.column_dimensions["B"].width = 45

    wb.save(path)


def save_removed_to_manual(removed):
    """把 hvoy 删掉的站写入 manual_sites.csv，确保后续脚本继续监测它们。"""
    if not removed:
        return

    # 读取已有 manual 条目，避免重复写入
    existing_domains = set()
    manual_exists = MANUAL_CSV.exists()
    fieldnames = [
        "source", "platform_name", "domain", "tech_stack", "favicon_group",
        "icp_filing", "has_privacy_policy", "contact_telegram", "contact_qq", "notes",
    ]
    if manual_exists:
        with open(MANUAL_CSV, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            if reader.fieldnames:
                fieldnames = reader.fieldnames
            for row in reader:
                existing_domains.add(row.get("domain", "").strip().lower())

    to_add = [
        p for p in removed
        if p.get("siteDomain", "").strip().lower() not in existing_domains
    ]

    if not to_add:
        print("  (已在 manual_sites.csv 中，无需重复写入)")
        return

    with open(MANUAL_CSV, "a", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        # 文件是新建的话要写表头
        if not manual_exists or MANUAL_CSV.stat().st_size == 0:
            writer.writeheader()
        for p in to_add:
            row = {field: "" for field in fieldnames}
            row.update({
                "source":        "hvoy_removed",
                "platform_name": p.get("siteName", ""),
                "domain":        p.get("siteDomain", ""),
            })
            writer.writerow(row)

    print(f"  已将 {len(to_add)} 个消失平台写入 manual_sites.csv（source=hvoy_removed）")


def make_path(directory, stem, suffix):
    p = directory / f"{stem}.{suffix}"
    if not p.exists():
        return p
    n = 2
    while True:
        p = directory / f"{stem}_{n}.{suffix}"
        if not p.exists():
            return p
        n += 1


def print_diff(added, removed):
    if added:
        print(f"\n🆕 新增 {len(added)} 个平台:")
        for p in added:
            print(f"  [{p['rankPosition']}] {p['siteName']} ({p['siteDomain']})")
    else:
        print("\n✅ 无新增平台")
    if removed:
        print(f"\n❌ 消失 {len(removed)} 个平台:")
        for p in removed:
            print(f"  {p['siteName']} ({p['siteDomain']})")
    else:
        print("✅ 无消失平台")


def main():
    HVOY_DIR.mkdir(parents=True, exist_ok=True)

    date_str    = datetime.now().strftime("%Y-%m-%d")
    latest_json = DATA_DIR / "hvoy_latest.json"
    latest_csv  = DATA_DIR / "hvoy_latest.csv"

    old_data = load_last(latest_json)
    try:
        items, updated_at = fetch_payload()
        save_fetch_status("SUCCESS", item_count=len(items))
    except Exception as exc:
        if not old_data or not latest_csv.exists():
            save_fetch_status("FAILED_NO_CACHE", error=exc)
            raise

        save_fetch_status(
            "FALLBACK_TO_CACHE",
            item_count=len(old_data),
            error=exc,
            used_cache=True,
        )
        print(f"::warning::hvoy 抓取失败，继续使用最近缓存（{len(old_data)}个平台）: {exc}")
        print(f"缓存文件: {latest_csv}")
        return

    added, removed = compare(old_data, items)

    print_diff(added, removed)

    # 把 hvoy 删掉的站落入 manual_sites.csv，让其他脚本继续监测
    if removed:
        print(f"\n📌 将消失平台写入 manual_sites.csv ...")
        save_removed_to_manual(removed)

    today_json = make_path(HVOY_DIR, f"hvoy_{date_str}", "json")
    today_xlsx = make_path(HVOY_DIR, f"hvoy_{date_str}", "xlsx")

    save_json(items, today_json)
    save_json(items, latest_json)
    save_csv(items, latest_csv)
    save_excel(items, added, removed, today_xlsx, updated_at, date_str)

    if added or removed:
        diff_file = make_path(HVOY_DIR, f"diff_{date_str}", "json")
        save_json({"date": date_str, "added": added, "removed": removed}, diff_file)
        print(f"\n📝 diff已保存: {diff_file}")

    print(f"\n💾 已保存:")
    print(f"   {today_json}")
    print(f"   {today_xlsx}")
    print(f"   {latest_csv}")
    print(f"\n   总平台数: {len(items)}  |  新增: {len(added)}  |  消失: {len(removed)}")


if __name__ == "__main__":
    main()
