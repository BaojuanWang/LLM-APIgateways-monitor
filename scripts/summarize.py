"""
存活状态汇总脚本
运行位置: scripts/summarize.py
读取 results/monitor_results.csv，输出每个平台的综合存活状态
"""

import csv
from collections import Counter, defaultdict
from datetime import datetime
from html import escape
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

BASE_DIR = Path(__file__).parent.parent
RESULTS_CSV = BASE_DIR / "results" / "monitor_results.csv"
OUTPUT_DIR = BASE_DIR / "results"
REVIEW_CSV = OUTPUT_DIR / "needs_review.csv"
STATUS_SVG = OUTPUT_DIR / "status_summary.svg"
DAILY_STATUS_DIR = OUTPUT_DIR / "daily"

ALIVE_STATUSES = {
    "ONLINE",
    "CLOUDFLARE_OR_BLOCKED",
    "ONLINE_LOGIN_REQUIRED",
    "HTTP_444",
    "ALIVE_BLOCKED",
    "REDIRECTED",
}
UNCERTAIN_STATUSES = {"TIMEOUT", "HTTP_ERROR"}
SERVICE_STOPPED_STATUSES = {"SERVICE_STOPPED"}
DEAD_STATUSES = {"DNS_FAIL", "PARKED_OR_FOR_SALE"} | SERVICE_STOPPED_STATUSES


def classify_overall(status_counts):
    statuses = set(status_counts.keys())
    if statuses & SERVICE_STOPPED_STATUSES:
        return "DEAD"
    if statuses & ALIVE_STATUSES:
        return "ALIVE"
    if statuses <= UNCERTAIN_STATUSES:
        return "UNCERTAIN"
    if statuses <= DEAD_STATUSES:
        return "DEAD"
    return "UNCERTAIN"


def status_bucket(status):
    if status in ALIVE_STATUSES:
        return "ALIVE"
    if status in DEAD_STATUSES:
        return "DEAD"
    return "UNCERTAIN"


def load_results():
    if not RESULTS_CSV.exists():
        raise FileNotFoundError(f"找不到 {RESULTS_CSV}")

    platforms = defaultdict(lambda: {
        "platform_name": "",
        "source": "",
        "source_category": "",
        "checks": [],
        "status_counts": defaultdict(int),
        "last_status": "",
        "last_timestamp": "",
        "last_final_url": "",
        "last_page_title": "",
    })

    with open(RESULTS_CSV, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            domain = row.get("domain", "").strip()
            if not domain:
                continue

            p = platforms[domain]
            p["platform_name"] = row.get("platform_name", "") or p["platform_name"]
            source = row.get("source", "")
            p["source"] = source
            p["source_category"] = "manual" if source in ("Taobao", "Xiaohongshu", "manual_stopped") else "hvoy"

            status = row.get("online_status", "")
            timestamp = row.get("timestamp", "")
            p["status_counts"][status] += 1
            p["checks"].append({"timestamp": timestamp, "status": status})

            if timestamp > p["last_timestamp"]:
                p["last_timestamp"] = timestamp
                p["last_status"] = status
                p["last_final_url"] = row.get("final_url", "")
                p["last_page_title"] = row.get("page_title", "")

    return platforms


def build_summary_rows(platforms):
    rows = []
    for domain, p in platforms.items():
        sc = p["status_counts"]
        overall = classify_overall(sc)
        known_statuses = ALIVE_STATUSES | DEAD_STATUSES | UNCERTAIN_STATUSES
        rows.append({
            "overall": overall,
            "domain": domain,
            "platform": p,
            "total": sum(sc.values()),
            "alive_n": sum(v for k, v in sc.items() if k in ALIVE_STATUSES),
            "uncertain_n": sum(v for k, v in sc.items() if k in UNCERTAIN_STATUSES or k not in known_statuses),
            "dead_n": sum(v for k, v in sc.items() if k in DEAD_STATUSES),
            "detail": ", ".join(f"{k}×{v}" for k, v in sorted(sc.items())),
        })

    order = {"DEAD": 0, "UNCERTAIN": 1, "ALIVE": 2}
    rows.sort(key=lambda x: (order.get(x["overall"], 9), x["domain"]))
    return rows


def make_date_path(stem, suffix):
    date_str = datetime.now().strftime("%Y-%m-%d")
    p = OUTPUT_DIR / f"{stem}_{date_str}.{suffix}"
    if not p.exists():
        return p
    n = 2
    while True:
        p = OUTPUT_DIR / f"{stem}_{date_str}_{n}.{suffix}"
        if not p.exists():
            return p
        n += 1


def save_needs_review(platforms):
    fields = [
        "domain", "platform_name", "综合判断", "检测次数",
        "各状态明细", "最近状态", "最近检测时间", "最终URL", "备注",
    ]
    rows = []
    for item in build_summary_rows(platforms):
        if item["overall"] != "UNCERTAIN":
            continue
        p = item["platform"]
        rows.append({
            "domain": item["domain"],
            "platform_name": p["platform_name"],
            "综合判断": item["overall"],
            "检测次数": item["total"],
            "各状态明细": item["detail"],
            "最近状态": p["last_status"],
            "最近检测时间": p["last_timestamp"],
            "最终URL": p["last_final_url"],
            "备注": "请用国内网络手动验证",
        })

    rows.sort(key=lambda x: x["domain"])
    with open(REVIEW_CSV, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)
    return len(rows)


def save_excel(platforms):
    wb = Workbook()
    rows = build_summary_rows(platforms)

    header_fill = PatternFill("solid", start_color="2D5EA2")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    alive_fill = PatternFill("solid", start_color="D4EFDF")
    uncertain_fill = PatternFill("solid", start_color="FEF9E7")
    dead_fill = PatternFill("solid", start_color="FADBD8")
    alt_fill = PatternFill("solid", start_color="F0F4FA")
    status_fill = {"ALIVE": alive_fill, "UNCERTAIN": uncertain_fill, "DEAD": dead_fill}

    headers = [
        "域名", "平台名称", "来源", "来源分类", "综合判断", "检测次数",
        "最近状态", "最近检测时间", "ALIVE次数", "UNCERTAIN次数", "DEAD次数",
        "各状态明细", "最终URL", "页面标题",
    ]
    col_widths = [28, 18, 12, 10, 10, 8, 24, 22, 8, 10, 8, 35, 35, 40]

    def write_header(ws):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"

    def row_values(item):
        p = item["platform"]
        return [
            item["domain"], p["platform_name"], p["source"], p["source_category"],
            item["overall"], item["total"], p["last_status"], p["last_timestamp"],
            item["alive_n"], item["uncertain_n"], item["dead_n"], item["detail"],
            p["last_final_url"], p["last_page_title"],
        ]

    def write_rows(ws, sheet_rows, forced_fill=None):
        write_header(ws)
        for row_idx, item in enumerate(sheet_rows, 2):
            fill = forced_fill or status_fill.get(item["overall"], alt_fill)
            for col, val in enumerate(row_values(item), 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(vertical="center")
                cell.fill = fill

    ws = wb.active
    ws.title = "综合汇总"
    write_rows(ws, rows)

    ws_dead = wb.create_sheet("疑似关闭")
    write_rows(ws_dead, [r for r in rows if r["overall"] == "DEAD"], dead_fill)

    ws_uncertain = wb.create_sheet("待确认")
    write_rows(ws_uncertain, [r for r in rows if r["overall"] == "UNCERTAIN"], uncertain_fill)

    total_platforms = len(platforms)
    alive_count = sum(1 for r in rows if r["overall"] == "ALIVE")
    uncertain_count = sum(1 for r in rows if r["overall"] == "UNCERTAIN")
    dead_count = sum(1 for r in rows if r["overall"] == "DEAD")

    ws_info = wb.create_sheet("说明")
    meta = [
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("数据来源", str(RESULTS_CSV)),
        ("平台总数", total_platforms),
        ("ALIVE", f"{alive_count} 个（入口确认存活）"),
        ("UNCERTAIN", f"{uncertain_count} 个（需进一步确认）"),
        ("DEAD", f"{dead_count} 个（DNS失败/域名停放/人工确认停止维护）"),
        ("", ""),
        ("判断规则", ""),
        ("ALIVE", "至少一次检测返回 ONLINE / Cloudflare / 444 / 登录页"),
        ("UNCERTAIN", "从未确认存活，且最近/历史结果仍需人工判断"),
        ("DEAD", "所有检测均为 DNS_FAIL / PARKED_OR_FOR_SALE，或人工确认 SERVICE_STOPPED"),
        ("最新概览图", str(STATUS_SVG)),
        ("每日概览图目录", str(DAILY_STATUS_DIR)),
    ]
    for r, (k, v) in enumerate(meta, 1):
        ws_info.cell(row=r, column=1, value=k).font = Font(bold=True, name="Arial")
        ws_info.cell(row=r, column=2, value=str(v)).font = Font(name="Arial")
    ws_info.column_dimensions["A"].width = 16
    ws_info.column_dimensions["B"].width = 64

    path = make_date_path("summary", "xlsx")
    wb.save(path)
    return path, alive_count, uncertain_count, dead_count


def latest_status_counts(platforms):
    return Counter(p["last_status"] or "UNKNOWN" for p in platforms.values())


def changed_since_previous_check(platforms):
    changes = []
    for domain, p in platforms.items():
        checks = sorted(
            [c for c in p["checks"] if c.get("timestamp")],
            key=lambda c: c["timestamp"],
        )
        if len(checks) < 2:
            continue
        previous = checks[-2]
        current = checks[-1]
        if previous["status"] == current["status"]:
            continue
        changes.append({
            "domain": domain,
            "platform_name": p["platform_name"],
            "previous": previous["status"],
            "current": current["status"],
            "timestamp": current["timestamp"],
        })
    changes.sort(key=lambda x: (status_bucket(x["current"]), x["domain"]))
    return changes


def svg_text(x, y, text, size=16, weight="400", fill="#111827"):
    return (
        f'<text x="{x}" y="{y}" font-family="Arial, sans-serif" '
        f'font-size="{size}" font-weight="{weight}" fill="{fill}">{escape(str(text))}</text>'
    )


def svg_rect(x, y, width, height, fill, radius=8, stroke="none"):
    return (
        f'<rect x="{x}" y="{y}" width="{width}" height="{height}" '
        f'rx="{radius}" fill="{fill}" stroke="{stroke}" />'
    )


def save_summary_plot(platforms, alive, uncertain, dead, review_count):
    """Write latest SVG plus one daily archive SVG for quick inspection in GitHub."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    DAILY_STATUS_DIR.mkdir(parents=True, exist_ok=True)

    total = max(alive + uncertain + dead, 1)
    rows = build_summary_rows(platforms)
    latest_counts = latest_status_counts(platforms)
    changes = changed_since_previous_check(platforms)
    latest_timestamp = max((p["last_timestamp"] for p in platforms.values()), default="")

    width = 1100
    height = 760
    chart_x = 70
    chart_w = 680
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    today = datetime.now().strftime("%Y-%m-%d")

    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
        svg_rect(0, 0, width, height, "#F8FAFC", radius=0),
        svg_text(60, 32, "LLM Relay Monitor - Status Overview", 28, "700"),
        svg_text(60, 62, f"Generated: {generated_at} | Latest data: {latest_timestamp or 'n/a'}", 14, "400", "#475569"),
    ]

    cards = [
        ("ALIVE", alive, "#16A34A", "Confirmed reachable at least once"),
        ("UNCERTAIN", uncertain, "#D97706", "Needs manual confirmation"),
        ("DEAD", dead, "#DC2626", "DNS failed, parked, or service stopped"),
    ]
    for i, (label, value, color, note) in enumerate(cards):
        x = 60 + i * 330
        parts.extend([
            svg_rect(x, 92, 300, 110, "#FFFFFF", stroke="#E2E8F0"),
            svg_text(x + 22, 125, label, 16, "700", color),
            svg_text(x + 22, 165, value, 36, "700", "#0F172A"),
            svg_text(x + 94, 163, f"/ {total}", 15, "400", "#64748B"),
            svg_text(x + 22, 188, note, 13, "400", "#64748B"),
        ])

    parts.append(svg_text(60, 248, "Overall classification", 18, "700"))
    bar_y = 272
    x = chart_x
    for label, value, color, _ in cards:
        segment_w = int(chart_w * value / total)
        if segment_w <= 0 and value > 0:
            segment_w = 2
        parts.append(svg_rect(x, bar_y, segment_w, 34, color, radius=4))
        if segment_w > 60:
            parts.append(svg_text(x + 10, bar_y + 23, f"{label} {value}", 13, "700", "#FFFFFF"))
        x += segment_w
    parts.append(svg_rect(chart_x, bar_y, chart_w, 34, "none", radius=4, stroke="#CBD5E1"))

    parts.append(svg_text(60, 352, "Latest check status distribution", 18, "700"))
    latest_total = max(sum(latest_counts.values()), 1)
    status_colors = {
        "ONLINE": "#16A34A",
        "CLOUDFLARE_OR_BLOCKED": "#22C55E",
        "ONLINE_LOGIN_REQUIRED": "#65A30D",
        "HTTP_404": "#D97706",
        "HTTP_ERROR": "#F59E0B",
        "TIMEOUT": "#FBBF24",
        "DNS_FAIL": "#DC2626",
        "PARKED_OR_FOR_SALE": "#B91C1C",
        "SERVICE_STOPPED": "#991B1B",
    }
    for idx, (status, count) in enumerate(latest_counts.most_common(8)):
        row_y = 382 + idx * 30
        bar_w = int(430 * count / latest_total)
        color = status_colors.get(status, "#64748B")
        parts.extend([
            svg_text(70, row_y + 18, status, 13, "700", "#334155"),
            svg_rect(275, row_y, max(bar_w, 2), 18, color, radius=4),
            svg_text(285 + max(bar_w, 2), row_y + 15, count, 13, "700", "#334155"),
        ])

    notes_x = 790
    parts.extend([
        svg_rect(notes_x, 238, 250, 416, "#FFFFFF", stroke="#E2E8F0"),
        svg_text(notes_x + 22, 272, "Notes", 18, "700"),
        svg_text(notes_x + 22, 306, f"Needs review: {review_count}", 14, "700", "#D97706"),
        svg_text(notes_x + 22, 330, f"Changed since previous check: {len(changes)}", 14, "700", "#334155"),
        svg_text(notes_x + 22, 364, "Review first", 14, "700", "#0F172A"),
    ])

    review_rows = [r for r in rows if r["overall"] in {"DEAD", "UNCERTAIN"}][:8]
    if not review_rows:
        parts.append(svg_text(notes_x + 22, 390, "No dead or uncertain platforms.", 13, "400", "#16A34A"))
    else:
        for idx, item in enumerate(review_rows):
            p = item["platform"]
            row_y = 390 + idx * 24
            label = f"{item['overall']}: {item['domain']} ({p['last_status']})"
            parts.append(svg_text(notes_x + 22, row_y, label[:38], 12, "400", "#475569"))

    parts.append(svg_text(notes_x + 22, 596, "Rules", 14, "700", "#0F172A"))
    rule_lines = [
        "ALIVE: any historical confirmed response.",
        "UNCERTAIN: no confirmed good response yet.",
        "DEAD: DNS fail, parked, or service stopped.",
    ]
    for idx, line in enumerate(rule_lines):
        parts.append(svg_text(notes_x + 22, 620 + idx * 20, line, 12, "400", "#64748B"))

    parts.append(svg_text(60, 676, "Recent status changes", 18, "700"))
    if not changes:
        parts.append(svg_text(70, 706, "No platform changed status compared with its previous check.", 13, "400", "#475569"))
    else:
        for idx, change in enumerate(changes[:6]):
            label = f"{change['domain']}: {change['previous']} -> {change['current']}"
            parts.append(svg_text(70, 706 + idx * 22, label, 13, "400", "#475569"))

    parts.append("</svg>")
    svg = "\n".join(parts)
    daily_svg = DAILY_STATUS_DIR / f"status_summary_{today}.svg"

    STATUS_SVG.write_text(svg, encoding="utf-8")
    daily_svg.write_text(svg, encoding="utf-8")
    return STATUS_SVG, daily_svg


def main():
    print("读取检测结果...")
    platforms = load_results()
    print(f"共 {len(platforms)} 个平台")

    path, alive, uncertain, dead = save_excel(platforms)
    review_count = save_needs_review(platforms)
    latest_plot_path, daily_plot_path = save_summary_plot(platforms, alive, uncertain, dead, review_count)

    print("\n── 综合判断 ──")
    print(f"  ALIVE     {alive}")
    print(f"  UNCERTAIN {uncertain}")
    print(f"  DEAD      {dead}")
    print(f"\n✅ 已保存: {path}")
    print(f"📊 最新概览图已覆盖: {latest_plot_path}")
    print(f"📊 每日概览图已覆盖: {daily_plot_path}")
    print(f"📋 需手动确认: {review_count} 个 → {REVIEW_CSV}")


if __name__ == "__main__":
    main()
